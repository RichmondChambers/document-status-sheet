import streamlit as st
import faiss
import pickle
import numpy as np
import re
import json
import requests
import jwt  # PyJWT
from openai import OpenAI
from pathlib import Path
import base64
import pandas as pd
import io
import string

from index_builder import sync_drive_and_rebuild_index_if_needed, INDEX_FILE, METADATA_FILE

# =========================
# Constants for feedback / corrections
# =========================

CORRECTIONS_FILE = "corrections.jsonl"
FEEDBACK_LOG_FILE = "feedback_log.jsonl"


# =========================
# 0. Google SSO
# =========================

def google_login():
    """
    Require the user to sign in with a Google account and restrict access
    to @richmondchambers.com email addresses.
    """
    if "user_email" in st.session_state:
        return st.session_state["user_email"]

    params = st.query_params
    if "code" in params:
        code = params["code"]
        token_response = requests.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": code,
                "client_id": st.secrets["GOOGLE_CLIENT_ID"],
                "client_secret": st.secrets["GOOGLE_CLIENT_SECRET"],
                "redirect_uri": st.secrets["GOOGLE_REDIRECT_URI"],
                "grant_type": "authorization_code",
            },
            timeout=15,
        )
        if token_response.status_code != 200:
            st.error("Authentication with Google failed. Please refresh and try again.")
            st.stop()

        token_data = token_response.json()
        id_token = token_data.get("id_token")
        if not id_token:
            st.error("No ID token received from Google.")
            st.stop()

        try:
            # NOTE: signature verification skipped for simplicity
            claims = jwt.decode(id_token, options={"verify_signature": False})
        except Exception:
            st.error("Could not decode ID token.")
            st.stop()

        email = claims.get("email", "")
        hosted_domain = claims.get("hd", "")

        if email.endswith("@richmondchambers.com") or hosted_domain == "richmondchambers.com":
            st.session_state["user_email"] = email
            return email
        else:
            st.error("Access is restricted to employees of Richmond Chambers.")
            st.stop()

    auth_url = (
        "https://accounts.google.com/o/oauth2/v2/auth"
        "?response_type=code"
        f"&client_id={st.secrets['GOOGLE_CLIENT_ID']}"
        f"&redirect_uri={st.secrets['GOOGLE_REDIRECT_URI']}"
        "&scope=openid%20email"
        "&prompt=select_account"
        "&access_type=offline"
    )

    st.markdown("### Richmond Chambers – Internal Tool")
    st.write(
        "Please sign in with a Richmond Chambers Google Workspace account to access this app."
    )
    st.markdown(f"[Sign in with Google]({auth_url})")
    st.stop()


# =========================
# 1. Keys + Auth (OpenAI client)
# =========================

client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
user_email = google_login()

# =========================
# 2. FAISS + metadata
# =========================

@st.cache_resource
def load_index_and_metadata():
    """
    Ensure FAISS index is up to date, then load index, metadata, and read last
    rebuilt timestamp for UI display.
    """
    sync_drive_and_rebuild_index_if_needed()
    index = faiss.read_index(INDEX_FILE)

    with open(METADATA_FILE, "rb") as f:
        metadata = pickle.load(f)

    try:
        with open("drive_index_state.json", "r") as f:
            state = json.load(f)
            last_rebuilt = state.get("last_rebuilt", "Unknown")
    except Exception:
        last_rebuilt = "Unknown"

    return index, metadata, last_rebuilt


index, metadata, last_rebuilt = load_index_and_metadata()

# =========================
# 3. File extraction
# =========================

def extract_text_from_uploaded_file(uploaded_file):
    name = uploaded_file.name.lower()

    if name.endswith(".txt"):
        return uploaded_file.read().decode("utf-8", errors="ignore")

    if name.endswith(".pdf"):
        try:
            import PyPDF2

            reader = PyPDF2.PdfReader(uploaded_file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            return text
        except Exception:
            return ""

    if name.endswith(".docx"):
        try:
            import docx

            doc = docx.Document(uploaded_file)
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            return ""

    try:
        return uploaded_file.read().decode("utf-8", errors="ignore")
    except Exception:
        return ""


# =========================
# 4. Embeddings + retrieval
# =========================

def get_embedding(text, model="text-embedding-3-small"):
    result = client.embeddings.create(input=[text], model=model)
    return result.data[0].embedding


def search_index(query, k=8, source_type=None):
    """
    Search FAISS for top-k chunks.
    source_type filters by metadata[i]["type"] in {"rule","precedent"}.
    """
    if not metadata:
        return []

    query_embedding = get_embedding(query)
    distances, indices_ = index.search(
        np.array([query_embedding], dtype=np.float32), k * 3
    )

    results = []
    for i in indices_[0]:
        if i < len(metadata):
            item = metadata[i]
            if source_type and item.get("type") != source_type:
                continue
            results.append(item)
        if len(results) >= k:
            break

    return results


# =========================
# 4A. Feedback / corrections helpers
# =========================

def append_jsonl(path: str, record: dict) -> None:
    """
    Append a single JSON record to a .jsonl file.
    """
    try:
        with open(path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception as e:
        # For internal use, surface as a warning rather than failing the app.
        st.warning(f"Could not write to {path}: {e}")


def load_corrections(path: str = CORRECTIONS_FILE) -> list[dict]:
    """
    Load all route-specific corrections from a JSONL file.
    Each record is expected to contain at least:
      - route_hint: str (lowercased substring of route description)
      - instruction: str (concise override instruction)
    """
    corrections: list[dict] = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    corrections.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    except FileNotFoundError:
        # No corrections yet
        pass
    except Exception as e:
        st.warning(f"Could not load corrections from {path}: {e}")
    return corrections


def summarise_feedback_to_instruction(route_text: str, feedback_text: str) -> str:
    """
    Use the OpenAI API to compress raw feedback into a single, precise
    route-specific instruction (under ~40 words) that corrects the error.
    Falls back to the original feedback if anything goes wrong.
    """
    route_clean = (route_text or "").strip()
    feedback_clean = (feedback_text or "").strip()

    if not feedback_clean:
        return ""

    try:
        resp = client.responses.create(
            model="gpt-5.1",
            input=[
                {
                    "role": "system",
                    "content": (
                        "You are a UK immigration lawyer. Rewrite the user's feedback into a single, precise instruction "
                        "that corrects an error in a document checklist. The sentence must be under 40 words, "
                        "route-specific, and phrased as a 'must' or 'must not' rule."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        f"ROUTE DESCRIPTION:\n{route_clean}\n\n"
                        f"FEEDBACK (legal error description):\n{feedback_clean}"
                    ),
                },
            ],
            temperature=0.0,
        )
        instruction = (getattr(resp, "output_text", "") or "").strip()
        if not instruction:
            instruction = feedback_clean
        return instruction
    except Exception as e:
        st.warning(f"Could not summarise feedback into an instruction: {e}")
        return feedback_clean


def get_applicable_correction_prompts(route_text: str) -> list[str]:
    """
    Return a list of correction instructions whose route_hint appears in the route text.
    Matching is deliberately simple and conservative.
    """
    if not route_text or not route_text.strip():
        return []

    route_lower = route_text.lower()
    prompts: list[str] = []
    for entry in load_corrections():
        hint = str(entry.get("route_hint", "")).lower().strip()
        # Support both 'instruction' and legacy 'prompt' keys
        instruction = str(entry.get("instruction") or entry.get("prompt") or "").strip()
        if hint and instruction and hint in route_lower:
            prompts.append(instruction)
    return prompts


def build_correction_prompt(route_text: str, feedback_text: str) -> dict | None:
    """
    Turn raw feedback into a route-scoped correction entry, using the model
    to create a concise one-line instruction.
    """
    route_hint = (route_text or "").strip()
    feedback_clean = (feedback_text or "").strip()

    if not route_hint or not feedback_clean:
        return None

    # Use a short route hint (first line / up to 120 chars), lowercased for matching
    first_line = route_hint.splitlines()[0].strip()
    if len(first_line) > 120:
        first_line = first_line[:117] + "..."
    route_hint_normalised = first_line.lower()

    concise_instruction = summarise_feedback_to_instruction(route_hint, feedback_clean)
    concise_instruction = concise_instruction.strip()
    if not concise_instruction:
        return None

    return {
        "route_hint": route_hint_normalised,
        "instruction": concise_instruction,
    }


def log_feedback_and_optionally_add_correction(
    user_email: str,
    route_text: str,
    facts_text: str,
    filter_label: str,
    tsv_output: str,
    feedback_text: str,
) -> None:
    """
    1. Log full feedback (for audit/training) to FEEDBACK_LOG_FILE.
    2. Add a lightweight route-specific correction entry to CORRECTIONS_FILE
       so future runs for similar routes pick up the override.
    """
    # 1. Full feedback log
    feedback_record = {
        "user_email": user_email,
        "route": route_text,
        "facts": facts_text,
        "filter_label": filter_label,
        "tsv_output": tsv_output,
        "feedback": feedback_text,
    }
    append_jsonl(FEEDBACK_LOG_FILE, feedback_record)

    # 2. Optional correction entry
    correction_entry = build_correction_prompt(route_text, feedback_text)
    if correction_entry is not None:
        append_jsonl(CORRECTIONS_FILE, correction_entry)


# =========================
# 5. Rule update date
# =========================

def fetch_latest_rule_update_date():
    """
    Light scrape of GOV.UK updates page.
    If it fails, fallback to today's date (UTC).
    """
    try:
        r = requests.get(
            "https://www.gov.uk/guidance/immigration-rules/updates", timeout=10
        )
        if r.status_code != 200:
            raise Exception("Bad status")

        m = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", r.text)
        if not m:
            raise Exception("No date found")
        return m.group(1)
    except Exception:
        return str(np.datetime64("today"))


# =========================
# 6. System prompt
# =========================

BASE_SYSTEM_PROMPT = """
You are a UK immigration lawyer specialising in document-checklist guidance for visa/immigration applications under the UK's Immigration Rules.

CRITICAL GROUNDING REQUIREMENT:
- You MUST use the lookup_rule tool to obtain the exact rule text supporting EVERY mandatory or recommended document requirement AND every discretion/judgment point.
- Do not invent citations. If you cannot find supporting text using lookup_rule, say so and limit guidance accordingly.

Hard rules:
- Base guidance ONLY on the Immigration Rules + official Home Office guidance. Avoid speculation.
- Quote/cite the exact relevant paragraph(s).
- Always reflect the April 2024 Appendix FM financial requirement (£29,000).
- Do NOT say two passport-sized photos are required.
- You MUST NOT produce a single combined document row for BRP and eVisa status.
- Instead, wherever current status and past BRPs are relevant, you MUST create TWO separate document rows:
  1) "Current Home Office eVisa status printout"
  2) "Previous BRPs"
- Do NOT refer to "Current BRP (if held)" in any document name. Assume status is now evidenced digitally via eVisa.
- If you are unsure whether a requirement applies, or you cannot find clear supporting rule text, you must treat that requirement as “Recommended (not mandatory)” and say so in Column G.
- You must not create maintenance / English language / accommodation requirements based solely on your general training; they must appear in the Rules or guidance you have been given.
- You MUST treat any “ROUTE-SPECIFIC CORRECTIONS / OVERRIDES” in the system/context as overriding your general knowledge for that route. Where there is a conflict, follow the corrections.
- Do NOT include a separate 'Legal Submissions', 'Cover Letter' or 'Advocacy Letter' document as a checklist item.

Section structure for a FULL DSS (no filter applied):
- You MUST organise the checklist into sections in this order and with these titles:
  * Section A: Application Documents
  * Section B: Nationality & Identity
  * Section C: Immigration History
  * Section D onwards: other logical sections required by the Immigration Rules (e.g. Financial Requirement, Relationship Requirement, Accommodation, English Language, Criminality, Suitability, etc.).
- Where there is a Relationship Requirement under the Immigration Rules or guidance, this MUST always have its own section (e.g. "Relationship Requirement" or "Relationship Evidence") and must NOT be merged into identity, nationality, immigration history or other sections.
- The final section in a full DSS MUST always be: "Document Format and Translations". This section should cover format requirements, translation requirements, copy certification, and any specific document-format rules.

Section structure for a FILTERED DSS (where a filter is requested):
- Begin with a Filtered Checklist row:
  Filtered Checklist: {filter_label}<TAB><TAB><TAB><TAB><TAB><TAB>
- Only include sections that are relevant to the filtered evidence type (for example, for "Accommodation evidence only", only accommodation-related sections and the document-format/translation section).
- The last section in a filtered DSS should still be a "Document Format and Translations" section, limited to the aspects relevant to the filtered documents.

Other output rules (SPREADSHEET-READY TSV, CLIENT + LAWYER REVIEW):
- You MUST output ONLY tab-separated values (TSV). No Markdown, no bullets, no numbering, no pipes "|".
- Produce EXACTLY 7 columns per row:
  Column A: Document
  Column B: Evidential Requirements (what must be shown)
  Column C: Client Notes (clear, client-ready explanation; no legalese)
  Column D: GDrive Link (leave blank)
  Column E: Ready To Review (leave blank; app will insert checkboxes)
  Column F: Status (leave blank; app will insert dropdown in Excel)
  Column G: Rule Authority (pinpoint paragraph reference + a SHORT supporting quotation)
- First row MUST be the header exactly:
  Document<TAB>Evidential Requirements<TAB>Client Notes<TAB>GDrive Link<TAB>Ready To Review<TAB>Status<TAB>Rule Authority
- Put section titles as a standalone row in Column A only, like:
  Section: Application Documents<TAB><TAB><TAB><TAB><TAB><TAB>
  Section: Nationality & Identity<TAB><TAB><TAB><TAB><TAB><TAB>
  Section: Immigration History<TAB><TAB><TAB><TAB><TAB><TAB>
  Section: Financial Requirement<TAB><TAB><TAB><TAB><TAB><TAB>
  etc.
  (Do NOT include any === symbols.)
- Call the main advocacy material simply part of the evidence within the relevant section (do not create a standalone Legal Submissions document row).
- NEVER start any cell with "=", "+", "-", or "@".
  If unavoidable, prefix that cell with "'".
- Client Notes must be concise, readable, and suitable to send to a client.
- Rule Authority must include:
  (i) the rule/appendix + paragraph reference, and
  (ii) a supporting quote no longer than ~25 words.
- If a filter is requested:
  * output ONLY that subset;
  * begin with a Filtered Checklist row as described above.
- If you cannot find supporting rule text using lookup_rule, state that briefly in Rule Authority
  and keep the document as "Recommended (not mandatory)" where appropriate.

Legal Authority Summary:
- Do NOT include a separate summary section. All authority must be in Column G only.
"""

# =========================
# 7. Tool implementation: lookup_rule
# =========================

def lookup_rule_tool(appendix_or_part, paragraph_ref=None, query=None):
    """
    Backend tool the model calls.
    We search RULE chunks only.
    """
    q = " ".join([s for s in [appendix_or_part, paragraph_ref, query] if s])
    hits = search_index(q, k=3, source_type="rule")

    inferred_ref = None
    if hits:
        refs = hits[0].get("paragraph_refs") or []
        inferred_ref = refs[0] if refs else None

    return {
        "appendix_or_part": appendix_or_part,
        "paragraph_ref": paragraph_ref or inferred_ref,
        "passages": [
            {
                "text": h.get("content", ""),
                "source": h.get("source", ""),
                "paragraph_ref": (h.get("paragraph_refs") or [""])[0],
            }
            for h in hits
        ],
    }


# =========================
# 8. Model call with tool loop (HARDENED)
# =========================

def generate_checklist(route_text, facts_text, extra_route_facts_text=None,
                       filter_mode=None, filter_label=None):
    rule_date = fetch_latest_rule_update_date()
    system_prompt = BASE_SYSTEM_PROMPT.replace("{RULE_UPDATE_DATE}", rule_date)

    enquiry_text = f"ROUTE:\n{route_text.strip()}\n\nFACTS:\n{facts_text.strip()}"

    rule_chunks = search_index(enquiry_text, k=10, source_type="rule")
    precedent_chunks = search_index(enquiry_text, k=4, source_type="precedent")

    grounding_context = "AUTHORITATIVE IMMIGRATION RULES EXTRACTS:\n"
    grounding_context += "\n\n".join(
        [
            f"[R{i+1}] ({(rc.get('appendix_or_part') or '')}"
            f"{' ' + (rc.get('paragraph_refs') or [''])[0] if rc.get('paragraph_refs') else ''}"
            f" — {rc.get('source','')})\n{rc.get('content','')}"
            for i, rc in enumerate(rule_chunks)
        ]
    )

    grounding_context += "\n\nINTERNAL PRECEDENT EXTRACTS (style only, not authority):\n"
    grounding_context += "\n\n".join(
        [
            f"[P{i+1}] ({pc.get('source','')})\n{pc.get('content','')}"
            for i, pc in enumerate(precedent_chunks)
        ]
    )

    if extra_route_facts_text and extra_route_facts_text.strip():
        grounding_context += "\n\nUPLOADED ROUTE/FACTS DOCUMENT:\n"
        grounding_context += extra_route_facts_text.strip()

    # Route-specific corrections based on prior feedback
    correction_prompts = get_applicable_correction_prompts(route_text)
    if correction_prompts:
        grounding_context += "\n\nROUTE-SPECIFIC CORRECTIONS / OVERRIDES (from prior feedback):\n"
        for cp in correction_prompts:
            grounding_context += f"- {cp}\n"

    user_instruction = enquiry_text
    if filter_mode:
        user_instruction += (
            f"\n\nFILTER LABEL: {filter_label}\n"
            f"FILTER REQUEST: {filter_mode}"
        )

    tools = [
        {
            "type": "function",
            "name": "lookup_rule",
            "description": "Return exact Immigration Rules text for a given appendix/part and paragraph reference.",
            "parameters": {
                "type": "object",
                "properties": {
                    "appendix_or_part": {"type": "string"},
                    "paragraph_ref": {"type": "string"},
                    "query": {"type": "string"},
                },
                "required": ["appendix_or_part"],
            },
        }
    ]

    resp = client.responses.create(
        model="gpt-5.1",
        input=[
            {"role": "system", "content": system_prompt},
            {"role": "system", "content": grounding_context},
            {"role": "user", "content": user_instruction},
        ],
        tools=tools,
        tool_choice="auto",
        temperature=0.2,
    )

    pending_tool_calls = []
    for item in getattr(resp, "output", []) or []:
        if getattr(item, "type", None) in ("tool_call", "function_call") and getattr(
            item, "name", None
        ) == "lookup_rule":
            pending_tool_calls.append(item)

    if not pending_tool_calls:
        return (getattr(resp, "output_text", "") or "").strip()

    resolved_rules = []
    for tc in pending_tool_calls:
        raw_args = getattr(tc, "arguments", None) or "{}"
        try:
            args = json.loads(raw_args) if isinstance(raw_args, str) else (raw_args or {})
        except Exception:
            args = {}

        tool_result = lookup_rule_tool(
            appendix_or_part=args.get("appendix_or_part", ""),
            paragraph_ref=args.get("paragraph_ref"),
            query=args.get("query"),
        )
        resolved_rules.append(tool_result)

    resolved_context = grounding_context + "\n\nLOOKUP_RULE RESULTS (authoritative):\n"
    for i, tr in enumerate(resolved_rules):
        resolved_context += f"\n[LR{i+1}] {tr.get('appendix_or_part','')} {tr.get('paragraph_ref','')}\n"
        for p in tr.get("passages", []):
            resolved_context += (
                f"- ({p.get('source','')} {p.get('paragraph_ref','')}) {p.get('text','')}\n"
            )

    followup = client.responses.create(
        model="gpt-5.1",
        input=[
            {"role": "system", "content": system_prompt},
            {"role": "system", "content": resolved_context},
            {"role": "user", "content": user_instruction},
        ],
        temperature=0.2,
    )

    return (getattr(followup, "output_text", "") or "").strip()


# =========================
# 9. Streamlit UI helpers
# =========================

def render_logo():
    logo_path = Path(__file__).parent / "logo.png"
    if logo_path.exists():
        b64 = base64.b64encode(logo_path.read_bytes()).decode()
        st.markdown(
            f"""
            <div style="display: flex; justify-content: center; padding-bottom: 10px;">
                <img src="data:image/png;base64,{b64}" width="150">
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.warning(f"Logo not found at {logo_path}")


def sanitize_for_sheets(tsv_text: str) -> str:
    lines = []
    for raw_line in tsv_text.splitlines():
        # Normalise any literal "<TAB>" or "\t" text into real tab characters
        line = raw_line.replace("<TAB>", "\t").replace("\\t", "\t")
        # Remove any stray pipe characters
        line = line.replace("|", "")

        # Remove trailing tab-created empty columns
        cols = line.split("\t")
        while cols and cols[-1] == "":
            cols.pop()

        # If more than 7 cols, merge middle into Client Notes (col C)
        if len(cols) > 7:
            head = cols[:2]       # Document, Evidential
            tail = cols[-4:]      # GDrive, Ready, Status, Authority
            notes_overflow = cols[2:-4]
            merged_notes = " ".join(c.strip() for c in notes_overflow if c.strip())
            cols = head + [merged_notes] + tail

        if len(cols) < 7:
            cols = cols + [""] * (7 - len(cols))
        elif len(cols) > 7:
            cols = cols[:6] + [" ".join(cols[6:])]

        safe_cols = []
        for c in cols:
            c = c.strip()
            c = c.replace("<TAB>", " ").replace("\\t", " ")
            if c.startswith(("=", "+", "-", "@")):
                c = "'" + c
            safe_cols.append(c)

        lines.append("\t".join(safe_cols))

    return "\n".join(lines).strip()

def remove_duplicate_header_rows(tsv_text: str) -> str:
    """
    Some model outputs repeat the header row as a first data row.
    This removes any non-first row that is effectively identical
    to the header (same columns after trimming).
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    # Normalised header
    header_cols = [c.strip() for c in lines[0].split("\t")]
    header_norm = "\t".join(header_cols)

    out = [lines[0]]  # keep the first header row

    for line in lines[1:]:
        cols = [c.strip() for c in line.split("\t")]
        norm = "\t".join(cols)
        if norm == header_norm:
            # skip duplicate header-as-row
            continue
        out.append(line)

    return "\n".join(out).strip()

def is_section_heading_row(cells: list[str]) -> bool:
    normalized = [c.lstrip("'").strip() for c in cells if c.strip()]
    if not normalized:
        return False
    first = normalized[0]
    return (
        first.startswith("Section")
        or first.startswith("Filtered Checklist")
        or first.startswith("===")
        or first.startswith("Section:")
    )


def standardise_document_names(tsv_text: str) -> str:
    """
    Normalise certain document names, e.g. legacy 'Cover Letter' names.
    (We later strip Legal Submissions / cover letters entirely.)
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []
    for i, line in enumerate(lines):
        cols = line.split("\t")
        if len(cols) < 7:
            cols = cols + [""] * (7 - len(cols))
        elif len(cols) > 7:
            cols = cols[:6] + [" ".join(cols[6:])]

        # Header row untouched
        if i == 0:
            out.append("\t".join(cols))
            continue

        doc_cell = cols[0].lstrip("'").strip().lower()

        # Only change true document rows, not section headings
        if doc_cell and not is_section_heading_row([cols[0]]):
            if "cover letter" in doc_cell or "covering letter" in doc_cell:
                cols[0] = "Legal Submissions"

        out.append("\t".join(cols))

    return "\n".join(out).strip()


def split_brp_evisa_rows(tsv_text: str) -> str:
    """
    If the model has produced a single combined BRP/eVisa document row
    (e.g. 'Current BRP (if held) and/or Home Office eVisa status printout'),
    replace that *one* row with TWO rows:
      - 'Current Home Office eVisa status printout'
      - 'Previous BRPs'
    Reuses the same columns B–G for both rows.
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []
    # Keep header as-is
    out.append(lines[0])

    for line in lines[1:]:
        cols = line.split("\t")
        if len(cols) < 7:
            cols = cols + [""] * (7 - len(cols))
        elif len(cols) > 7:
            cols = cols[:6] + [" ".join(cols[6:])]

        doc_raw = cols[0]
        doc_cell = doc_raw.lstrip("'").strip().lower()

        # Leave section headings etc. alone
        if is_section_heading_row([doc_raw]):
            out.append("\t".join(cols))
            continue

        # Detect a combined BRP + eVisa line
        has_brp = "brp" in doc_cell
        has_evisa = any(
            token in doc_cell
            for token in ["evisa", "e-visa", "e visa"]
        )

        if has_brp and has_evisa:
            # Use same evidential requirements / notes / authority for both rows
            common_cols = cols[1:]  # B–G

            row_evisa = ["Current Home Office eVisa status printout"] + common_cols
            row_brp = ["Previous BRPs"] + common_cols

            out.append("\t".join(row_evisa))
            out.append("\t".join(row_brp))
        else:
            out.append("\t".join(cols))

    return "\n".join(out).strip()


# =========================
# Witness Statement helpers (Section A)
# =========================

ORDINAL_LABELS = {
    1: "First Applicant",
    2: "Second Applicant",
    3: "Third Applicant",
    4: "Fourth Applicant",
    5: "Fifth Applicant",
}


def infer_applicant_labels(route_text: str, facts_text: str) -> list[str]:
    """
    Very simple heuristic:
    - Look for 'first applicant', 'second applicant', etc.
    - If not found, look for 'applicant 1', 'applicant 2', etc.
    - If still not found but 'applicant' appears, create generic
      First/Second/Third Applicant labels.
    - If nothing at all, assume a single Applicant.
    """
    blob = f"{route_text}\n{facts_text}"
    lower_blob = blob.lower()
    labels: list[str] = []
    seen_indices: set[int] = set()

    # 1. Explicit 'first applicant', 'second applicant', etc.
    for idx, word in [(1, "first"), (2, "second"), (3, "third"),
                      (4, "fourth"), (5, "fifth")]:
        if re.search(rf"\b{word}\s+applicant\b", lower_blob):
            labels.append(f"{word.capitalize()} Applicant")
            seen_indices.add(idx)

    # 2. 'applicant 1', 'applicant 2', etc.
    for m in re.findall(r"\bapplicant\s+(\d+)\b", lower_blob):
        try:
            idx = int(m)
        except ValueError:
            continue
        if 1 <= idx <= 10 and idx not in seen_indices:
            seen_indices.add(idx)
            if idx in ORDINAL_LABELS:
                labels.append(ORDINAL_LABELS[idx])
            else:
                labels.append(f"Applicant {idx}")

    # 3. If no structured labels, infer from number of 'applicant' mentions
    if not labels:
        count_mentions = len(re.findall(r"\bapplicants?\b", lower_blob))
        if count_mentions > 1:
            # Cap at 5 ordinals for sanity
            for i in range(1, min(count_mentions, 5) + 1):
                if i in ORDINAL_LABELS:
                    labels.append(ORDINAL_LABELS[i])
                else:
                    labels.append(f"Applicant {i}")
        elif count_mentions == 1:
            labels.append("Applicant")

    # 4. Fallback – single Applicant
    if not labels:
        labels.append("Applicant")

    return labels


def ensure_witness_statements_in_section_a(
    tsv_text: str, route_text: str, facts_text: str
) -> str:
    """
    Ensure Section A / Application Documents contains a witness statement
    row for each applicant inferred from the user text, and that those rows
    appear as the *last* document rows in that section.

    Works on 7-column TSV (pre-numbering).
    Behaviour:
    - Find the Application Documents section.
    - Remove any existing witness-statement rows in that section.
    - Append one witness statement row per inferred applicant *before*
      the next section heading (or end of file).
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    # 1. Locate the Application Documents section
    section_start = None
    section_end = len(lines)

    for i, line in enumerate(lines):
        cols = line.split("\t")
        if len(cols) < 7:
            cols += [""] * (7 - len(cols))
        heading = cols[0].lstrip("'").strip().lower()

        # Fuzzy match – before relettering it will usually be
        # "Section: Application Documents" or similar
        if heading.startswith("section") and "application" in heading and "document" in heading:
            section_start = i
            break

    if section_start is None:
        # No Application Documents section found
        return tsv_text

    # Find the next section heading after Application Documents
    for j in range(section_start + 1, len(lines)):
        c0 = lines[j].split("\t")[0].lstrip("'").strip().lower()
        if c0.startswith("section"):
            section_end = j
            break

    # 2. Build prefix: keep everything up to the end of Section A,
    # but strip any existing witness statement rows inside that section
    new_lines = []

    # Everything before the section heading stays as-is
    new_lines.extend(lines[: section_start + 1])

    # Inside Section A: keep non–witness-statement rows only
    for k in range(section_start + 1, section_end):
        c0 = lines[k].split("\t")[0].lstrip("'").strip().lower()
        if "witness statement" in c0:
            # Drop existing witness statement rows in Section A
            continue
        new_lines.append(lines[k])

    # 3. Append new witness statement rows at the end of Section A
    labels = infer_applicant_labels(route_text, facts_text)
    for label in labels:
        doc_name = f"{label}'s Witness Statement"
        row = [
            doc_name,
            "Signed witness statement setting out relevant background, facts and explanations in support of the application.",
            "We will draft this with you. Please check the content carefully, address any matters highlighted in red and ensure it is accurate and complete.",
            "",
            "",
            "",
            "No specific rule requiring a witness statement; recommended advocacy document to explain facts and address discretion.",
        ]
        new_lines.append("\t".join(row))

    # 4. Append the rest of the DSS (sections after Application Documents)
    new_lines.extend(lines[section_end:])

    return "\n".join(new_lines)


def remove_legal_submissions_rows(tsv_text: str) -> str:
    """
    Remove any document rows that are effectively 'Legal Submissions' /
    'cover letter' / 'advocacy letter' from the DSS.

    - Leaves section headings and other rows untouched.
    - Works on 7-column TSV (pre-numbering).
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []
    # Keep header
    out.append(lines[0])

    for i, line in enumerate(lines[1:], start=1):
        cols = line.split("\t")
        if len(cols) < 7:
            cols += [""] * (7 - len(cols))

        doc_cell_raw = cols[0]
        doc_cell = doc_cell_raw.lstrip("'").strip().lower()

        # Keep section/filtered headings
        if is_section_heading_row([doc_cell_raw]):
            out.append(line)
            continue

        # Blank / non-document rows pass through
        if not doc_cell:
            out.append(line)
            continue

        # Detect legal submissions / cover letters / advocacy letters
        if (
            "legal submissions" in doc_cell
            or "cover letter" in doc_cell
            or "covering letter" in doc_cell
            or "advocacy letter" in doc_cell
        ):
            # Skip this row entirely
            continue

        out.append(line)

    return "\n".join(out).strip()


# =========================
# Section heading + layout helpers
# =========================

def reletter_section_headings(tsv_text: str) -> str:
    """
    Rewrite section heading rows to:
      Section A: Title
      Section B: Title
      ...
    Removes any ===.
    Leaves Filtered Checklist row unlettered.

    IMPORTANT: this function DOES NOT move any document rows between sections,
    and DOES NOT change the semantic title text beyond cleaning prefixes.
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []
    letter_idx = 0
    letters = list(string.ascii_uppercase)

    for i, line in enumerate(lines):
        cols = line.split("\t")
        if len(cols) < 7:
            cols = cols + [""] * (7 - len(cols))
        elif len(cols) > 7:
            cols = cols[:6] + [" ".join(cols[6:])]

        if i == 0:
            out.append("\t".join(cols))
            continue

        doc_cell = cols[0].lstrip("'").strip()

        # Leave "Filtered Checklist" heading without section lettering
        if doc_cell.lower().startswith("filtered checklist"):
            cleaned = re.sub(r"^=+\s*", "", doc_cell).strip()
            cleaned = cleaned.replace("<TAB>", " ").replace("\\t", " ")
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            cols[0] = cleaned
            cols[1:] = [""] * 6
            out.append("\t".join(cols))
            continue

        if is_section_heading_row([doc_cell]):
            # Base cleaning
            cleaned = doc_cell.replace("===", "")
            cleaned = cleaned.replace("<TAB>", " ").replace("\\t", " ")
            cleaned = re.sub(r"^Section\s*:?\s*", "", cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r"^Section\s+[A-Z]\s*:?\s*", "", cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r"\s+", " ", cleaned).strip()

            letter = (
                letters[letter_idx]
                if letter_idx < len(letters)
                else f"({letter_idx+1})"
            )
            letter_idx += 1

            cols[0] = f"Section {letter}: {cleaned}"
            cols[1:] = [""] * 6
            out.append("\t".join(cols))
            continue

        out.append("\t".join(cols))

    return "\n".join(out).strip()


def remove_blank_rows(tsv_text: str) -> str:
    out = []
    for i, line in enumerate(tsv_text.splitlines()):
        cols = [c.strip() for c in line.split("\t")]
        if i == 0:
            out.append(line)
            continue
        if any(c for c in cols):
            out.append(line)
    return "\n".join(out).strip()


def add_numbering_column(tsv_text: str) -> str:
    """
    Add leftmost numbering column ("No.") so only document rows are numbered.
    For section rows, move the heading into Column A (No.) and blank the rest.
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []
    counter = 0

    for i, line in enumerate(lines):
        cols = line.split("\t")
        if len(cols) < 7:
            cols = cols + [""] * (7 - len(cols))
        elif len(cols) > 7:
            cols = cols[:6] + [" ".join(cols[6:])]

        stripped_cols = [c.strip() for c in cols]
        doc_cell_norm = stripped_cols[0].lstrip("'").strip()

        if i == 0:
            out.append("\t".join(["No."] + stripped_cols))
            continue

        is_heading = is_section_heading_row([doc_cell_norm])
        is_blank_row = doc_cell_norm == ""
        is_true_document_row = (doc_cell_norm != "") and (not is_heading)

        if is_heading:
            heading_text = stripped_cols[0]
            out.append("\t".join([heading_text] + [""] * 7))
            continue

        if is_blank_row or not is_true_document_row:
            out.append("\t".join([""] + stripped_cols))
        else:
            counter += 1
            out.append("\t".join([str(counter)] + stripped_cols))

    return "\n".join(out).strip()


def add_ready_checkboxes(tsv_text: str) -> str:
    """
    Insert a checkbox marker "☐" into Ready To Review column for document rows only.
    (Sheets checkboxes will be applied via Apps Script after paste.)
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []
    is_numbered = False

    for i, line in enumerate(lines):
        cols = line.split("\t")

        if i == 0:
            is_numbered = cols[0].strip().lower() in ("no.", "no", "#")

        if is_numbered:
            if len(cols) < 8:
                cols = cols + [""] * (8 - len(cols))
            elif len(cols) > 8:
                cols = cols[:7] + [" ".join(cols[7:])]
            ready_idx = 5  # F in numbered view
            doc_idx = 1
            no_idx = 0
        else:
            if len(cols) < 7:
                cols = cols + [""] * (7 - len(cols))
            elif len(cols) > 7:
                cols = cols[:6] + [" ".join(cols[6:])]
            ready_idx = 4  # E in unnumbered view
            doc_idx = 0
            no_idx = None

        stripped_cols = [c.strip() for c in cols]

        if i == 0:
            out.append("\t".join(stripped_cols))
            continue

        doc_cell = stripped_cols[doc_idx].lstrip("'").strip()
        is_heading = is_section_heading_row([doc_cell]) or stripped_cols[0].lstrip("'").strip().startswith("Section")
        is_blank_row = doc_cell == ""
        is_document_row = (doc_cell != "") and (not is_heading) and (not is_blank_row)

        if is_document_row:
            if no_idx is not None:
                no_cell = stripped_cols[no_idx].strip()
                cols[ready_idx] = "☐" if no_cell else ""
            else:
                cols[ready_idx] = "☐"
        else:
            cols[ready_idx] = ""

        out.append("\t".join(c.strip() for c in cols))

    return "\n".join(out).strip()


def move_col_g_to_h(tsv_text: str) -> str:
    """
    For 8-column TSVs (numbered view):
    - leave header row unchanged
    - leave section/filtered-heading rows unchanged
    - move any text in column G (index 6) to column H (index 7) and blank column G
    """
    lines = tsv_text.splitlines()
    if not lines:
        return tsv_text

    out = []

    for i, line in enumerate(lines):
        cols = line.split("\t")

        # Header row stays as-is
        if i == 0:
            out.append(line)
            continue

        # Ensure 8 columns
        if len(cols) < 8:
            cols = cols + [""] * (8 - len(cols))
        elif len(cols) > 8:
            cols = cols[:7] + [" ".join(cols[7:])]

        # After numbering, headings sit in col A (No.)
        heading_cell = cols[0].lstrip("'").strip()
        is_heading = is_section_heading_row([heading_cell])

        if is_heading:
            out.append("\t".join(cols))
            continue

        g_idx, h_idx = 6, 7
        g_val = cols[g_idx].strip()
        h_val = cols[h_idx].strip()

        if g_val:
            cols[h_idx] = (h_val + " " + g_val).strip() if h_val else g_val
            cols[g_idx] = ""

        out.append("\t".join(cols))

    return "\n".join(out)


def tsv_to_dataframe(tsv_text: str) -> pd.DataFrame:
    if not tsv_text.strip():
        return pd.DataFrame()
    return pd.read_csv(io.StringIO(tsv_text), sep="\t", dtype=str).fillna("")


def dataframe_to_formatted_xlsx_bytes(df: pd.DataFrame, sheet_name="Status Sheet") -> bytes:
    """
    Export DataFrame to formatted Excel (openpyxl):
    - wrap text
    - centre align
    - bold header (font size 14)
    - freeze row 1
    - borders
    - section rows: brand blue #009fdf, white bold text (font size 14), merged across row
    - Status: Excel dropdown list for document rows
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise ImportError(
            "openpyxl is required for formatted Excel export. "
            "Add 'openpyxl' to requirements.txt."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    section_fill = PatternFill(
        fill_type="solid", start_color="FF009FDF", end_color="FF009FDF"
    )
    section_font = Font(bold=True, color="FFFFFF", size=14)
    section_alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )

    # Header
    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=14)
        cell.alignment = header_alignment
        cell.border = border

    # Body
    for row_vals in df.itertuples(index=False):
        ws.append(list(row_vals))

    # Find status column index
    status_col_idx = None
    for j, col_name in enumerate(df.columns, start=1):
        if str(col_name).strip().lower() == "status":
            status_col_idx = j
            break

    # Dropdown for Status (Excel)
    status_options = [
        "Pending",
        "Approved",
        "Amendment Required",
        "Continue Uploading",
        "No Longer Required",
    ]
    if status_col_idx is not None:
        formula = '"' + ",".join(status_options) + '"'
        dv_status = DataValidation(
            type="list", formula1=formula, allow_blank=True, showDropDown=True
        )
        ws.add_data_validation(dv_status)
    else:
        dv_status = None

    # Style + merge section rows + apply dropdown only to document rows
    for r_idx in range(2, ws.max_row + 1):
        row_values = [
            str(ws.cell(row=r_idx, column=c_idx).value or "").strip()
            for c_idx in range(1, ws.max_column + 1)
        ]
        is_section = any(
            v.startswith("Section ") or v.startswith("Filtered Checklist")
            for v in row_values
            if v
        )

        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            cell.alignment = cell_alignment

            if is_section:
                cell.font = section_font
                cell.fill = section_fill

        if is_section:
            ws.merge_cells(
                start_row=r_idx,
                start_column=1,
                end_row=r_idx,
                end_column=ws.max_column,
            )
            ws.cell(row=r_idx, column=1).alignment = section_alignment
        else:
            if status_col_idx is not None and dv_status is not None:
                first_header = str(ws.cell(row=1, column=1).value or "").strip().lower()
                is_numbered = first_header in ("no.", "no", "#")
                doc_col_idx = 2 if is_numbered else 1
                doc_val = str(
                    ws.cell(row=r_idx, column=doc_col_idx).value or ""
                ).strip()
                if doc_val:
                    col_letter = get_column_letter(status_col_idx)
                    dv_status.add(f"{col_letter}{r_idx}")

    ws.freeze_panes = "A2"

    # Column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)] + [str(v) for v in df.iloc[:, col_idx - 1].values]
        max_len = max(len(v) for v in values)
        width = min(max(max_len * 0.9, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# =========================
# Filter options (label -> precise model instruction)
# =========================

FILTER_OPTIONS = {
    "Full DSS": None,
    "Mandatory documents only": (
        "Return ONLY documents that are strictly mandatory under the Immigration Rules or Home Office guidance "
        "(i.e., specified evidence or items without which the application is likely to be refused/invalid). "
        "Exclude purely recommended/supporting documents unless they are needed to satisfy a mandatory requirement."
    ),
    "Identity / nationality evidence only": (
        "Return ONLY identity and nationality evidence. Include passports/travel documents, BRP/eVisa status proof, "
        "national ID cards, biometrics/photo requirements where relevant, name-change evidence, and any identity-linked "
        "civil documents. Exclude financial, relationship, accommodation, or other categories."
    ),
    "Immigration history / status evidence only": (
        "Return ONLY evidence relating to the applicant’s UK immigration history and current/previous status. "
        "Include current grant/leave evidence, visa vignettes, entry/exit stamps, prior approvals/refusals/curtailments, "
        "overstay or breach explanations, conditions of leave, and any required history disclosures."
    ),
    "Application forms / administrative evidence only": (
        "Return ONLY administrative/process documents needed to lodge the application. "
        "Include online application form confirmation, IHS/payment receipts, appointment/biometrics confirmation, "
        "consent forms, document checklists, and any required declarations. "
        "Include translation/certification requirements ONLY insofar as they relate to admin validity."
    ),
    "Financial requirement evidence only": (
        "Return ONLY evidence proving the financial requirement for the route. "
        "Include specified evidence per the Rules/guidance: payslips, bank statements, employer letters, tax returns, "
        "audited/unaudited accounts, dividend vouchers, savings evidence, pension evidence, benefits letters, "
        "and evidence of source/ownership of funds where required."
    ),
    "Accommodation evidence only": (
        "Return ONLY evidence proving adequate accommodation. "
        "Include tenancy/mortgage documents, landlord/freeholder consent, property ownership proof, "
        "property inspection/overcrowding reports where relevant, utilities/council tax if used to show residence, "
        "and sponsor residence proof tied to accommodation adequacy."
    ),
    "English language evidence only": (
        "Return ONLY evidence proving English language requirement or exemption. "
        "Include approved SELT certificate, degree taught in English plus ECCTIS/UK NARIC comparability if required, "
        "nationality-based exemptions, age/medical exemptions with supporting proof."
    ),
    "Relationship / family evidence only": (
        "Return ONLY evidence proving relationship/family link. "
        "Include marriage/civil partnership certificates, divorce/death certificates, evidence of durable partnership, "
        "cohabitation evidence, communication/interaction evidence, family composition proof. "
        "Exclude genuineness narratives unless needed to prove relationship."
    ),
    "Genuine relationship / genuine intention evidence only": (
        "Return ONLY evidence aimed at demonstrating a genuine relationship or genuine intention to live together/continue "
        "the relationship (where required). "
        "Include relationship timeline, communications, visits, shared life evidence, joint commitments, "
        "statements addressing genuineness."
    ),
    "Employment / role evidence only": (
        "Return ONLY evidence of a job/role where employment is a core requirement. "
        "Include CoS, contract, SOC code fit evidence, salary breakdown, start date confirmation, "
        "sponsor licence/status proof, duties letters."
    ),
    "Qualifications / skills evidence only": (
        "Return ONLY evidence of qualifications/skills relevant to eligibility. "
        "Include degree certificates/transcripts, professional qualifications, registrations/licences, "
        "ATAS if applicable."
    ),
    "Maintenance / funds evidence only (non-salary routes)": (
        "Return ONLY evidence of maintenance/funds where the route requires proof of funds rather than salary "
        "(e.g., Student/Visitor/PBS dependants). "
        "Include bank statements showing required level and holding period, proof of ownership/control."
    ),
    "Sponsor evidence only": (
        "Return ONLY evidence relating to the sponsor. "
    ),
    "Dependants’ evidence only": (
        "Return ONLY evidence needed for dependants."
    ),
    "Children / parental responsibility evidence only": (
        "Return ONLY evidence concerning children and parental responsibility."
    ),
    "Study / CAS evidence only": (
        "Return ONLY study-related evidence for Student routes."
    ),
    "Business / investment evidence only": (
        "Return ONLY business/investment-related evidence."
    ),
    "TB / medical evidence only": (
        "Return ONLY TB/medical evidence."
    ),
    "Criminality / character evidence only": (
        "Return ONLY evidence relating to criminality/character."
    ),
    "Suitability / refusal-risk evidence only": (
        "Return ONLY evidence aimed at addressing suitability or refusal risks."
    ),
    "Exceptional circumstances / discretion evidence only": (
        "Return ONLY evidence supporting exceptional circumstances or discretionary grants."
    ),
    "Translations / format / copy certification evidence only": (
        "Return ONLY evidence about translations, formatting, and certification."
    ),
    "Country-specific evidence only": (
        "Return ONLY country-specific evidence requirements triggered by nationality/location."
    ),
}

# =========================
# 10. Streamlit UI
# =========================

render_logo()
st.markdown(
    "<h1 style='text-align: center; font-size: 2.6rem;'>Document Status Sheet Generator</h1>",
    unsafe_allow_html=True,
)
st.markdown(
    f"<p style='color: grey; text-align: center; font-size: 0.9rem;'>"
    f"Immigration Rules index last rebuilt from Drive on: <b>{last_rebuilt}</b></p>",
    unsafe_allow_html=True,
)

st.markdown(
    "Provide the immigration route and relevant case facts in separate fields. "
    "The app will generate a DSS. "
)
st.info(
    "Please download as Excel. Then copy/paste into a copy of your Google Sheets DSS template. "
)

uploaded_doc = st.file_uploader(
    "Optional: upload a document describing the immigration route and relevant facts.",
    type=["pdf", "txt", "docx"],
    help="E.g., case summary, client instructions, or notes setting out the route and facts.",
)

filter_label = st.selectbox(
    "DSS filter (optional)", list(FILTER_OPTIONS.keys()), index=0
)
filter_instruction = FILTER_OPTIONS[filter_label]

with st.form("checklist_form"):
    route = st.text_area(
        "Immigration Route",
        height=140,
        placeholder=(
            "Provide the immigration route, type of application (EC, FLR, ILR etc) "
            "and Appendix of the Rules."
        ),
    )
    facts = st.text_area(
        "Relevant Facts",
        height=260,
        placeholder=(
            "Provide the key case facts needed to generate the DSS. "
            "The more case detail you provide, the more case specific the output will be. "
            "Plain English is fine."
        ),
    )
    submit = st.form_submit_button("Generate DSS")

if submit and (route.strip() or facts.strip()):
    with st.spinner("Retrieving Rules, checking precedents, and generating DSS..."):
        extra_text = None
        if uploaded_doc is not None:
            extra_text = extract_text_from_uploaded_file(uploaded_doc)

        reply = generate_checklist(
            route_text=route,
            facts_text=facts,
            extra_route_facts_text=extra_text,
            filter_mode=filter_instruction,
            filter_label=filter_label,
        )

        reply = sanitize_for_sheets(reply)
        reply = remove_duplicate_header_rows(reply)  # <-- NEW LINE
        reply = standardise_document_names(reply)
        # Split any combined BRP / eVisa row into two separate rows
        reply = split_brp_evisa_rows(reply)
        # Ensure witness statements are added as the last docs in Section A
        reply = ensure_witness_statements_in_section_a(reply, route, facts)
        # Remove Legal Submissions / cover letter / advocacy letter rows entirely
        reply = remove_legal_submissions_rows(reply)
        reply = reletter_section_headings(reply)
        reply = remove_blank_rows(reply)
        reply = add_numbering_column(reply)
        reply = add_ready_checkboxes(reply)
        reply = move_col_g_to_h(reply)

        # Store in session_state so it persists across reruns
        st.session_state["tsv"] = reply
        st.session_state["route"] = route
        st.session_state["facts"] = facts
        st.session_state["filter_label"] = filter_label

# Display preview, download, and feedback once a DSS exists
if "tsv" in st.session_state and st.session_state["tsv"].strip():
    tsv_output = st.session_state["tsv"]
    st.success("DSS generated.")

    st.subheader("DSS Preview")
    st.code(tsv_output, language="text")

    # =========================
    # Professional Responsibility Statement
    # (now appears immediately after the response / preview)
    # =========================
    st.markdown(
        """
**Professional Responsibility Statement**

AI-generated content must not be relied upon without human review. Where such content is used, the barrister is responsible for verifying and ensuring the accuracy and legal soundness of that content. AI tools are used solely to support drafting and research; they do not replace the barrister’s independent judgment, analysis, or duty of care.
""",
        unsafe_allow_html=False,
    )

    df = tsv_to_dataframe(tsv_output)
    if not df.empty:
        try:
            xlsx_bytes = dataframe_to_formatted_xlsx_bytes(
                df, sheet_name="Status Sheet"
            )
            st.download_button(
                "Download formatted Excel (.xlsx)",
                data=xlsx_bytes,
                file_name="document_status_sheet.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )
        except Exception as e:
            st.warning(f"Formatted Excel export unavailable: {e}")

    # =========================
    # Feedback UI (appears only after DSS generated)
    # Now *after* the download button
    # =========================

    st.markdown("---")
    st.subheader("Report a legal error")

    st.write(
        "If you spot any incorrect or missing legal requirements (for example, a maintenance funds requirement "
        "where none exists), please describe the issue below. Your feedback will be logged and used to refine "
        "future DSS's for similar routes."
    )

    feedback_text = st.text_area(
        "Describe the legal issue and the correct position:",
        height=160,
        placeholder=(
            "Example:\n"
            "For Tier 1 (Investor) extension applications there is no maintenance funds requirement. "
            "The DSS should not include a maintenance funds section for this route."
        ),
        key="feedback_text_area",
    )

    if st.button("Submit feedback on this DSS"):
        if feedback_text.strip():
            route_state = st.session_state.get("route", "")
            facts_state = st.session_state.get("facts", "")
            filter_state = st.session_state.get("filter_label", "")

            log_feedback_and_optionally_add_correction(
                user_email=user_email,
                route_text=route_state,
                facts_text=facts_state,
                filter_label=filter_state,
                tsv_output=tsv_output,
                feedback_text=feedback_text,
            )
            st.success(
                "Thank you. Your feedback has been recorded and a route-specific override has been added "
                "for future DSS generation for similar routes."
            )
            # Optionally clear the text area
            st.session_state["feedback_text_area"] = ""
        else:
            st.warning("Please enter some feedback before submitting.")


